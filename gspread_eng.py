import os
import gspread
import openpyxl
import datetime
from oauth2client.service_account import ServiceAccountCredentials

# 구글 문서 연결
def conn_gspred(SHEET_URL, month_no):
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/Hyun/Downloads/quickstart-1575028783267-e76b12a95dfe.json', SCOPES)
    gc = gspread.authorize(credentials)
    sh1 = gc.open_by_url(SHEET_URL)
    sheet_title = f'{month_no}월'
    ws = sh1.worksheet(sheet_title)
    list_all = ws.get_all_values()
    return list_all


def write_message(list_all, month_no, week_no):
    sel_list = []
    for row in list_all: # 주차 별 출석일이 비어있지 않은 리스트 sel_list 로 반환
        if row[4] == week_no and row[5] != '':
            sel_list.append(row)

    name_list = []
    for row in sel_list: # 리스트 중 이름 목록 추출
        name_list.append(row[1])

    result = []
    for a in name_list: # 이름 목록 중 중복 제거
        if result.count(a) < 1:
            result.append(a)
    
    return_list = []
    for name in result:
        temp_list = [] # 이름 따라 개별 추출
        for row in sel_list:
            if row[1] == name:
                temp_list.append(row)
        init_line = f"[영어 {month_no}월 {week_no}주차]\n<{name} 학생>\n"
        body=""
        comments = ''
        sum_list = []
        for i in range(0,8):
            sum_list.append(0)

        for i in temp_list:
            date = i[5]
            status = i[6]
            pre = i[15]
            rev = i[16]
            att = i[17]
            comm = ''
            if i[18] != '':
                comm = i[18]+'\n'
            item_tuple = i[7], i[8], i[9], i[10], i[11], i[12], i[13], i[14]
            item_list = []
            for i in item_tuple:
                if i != '':
                    try:
                        a = int(i)
                    except:
                        a = 0
                    item_list.append(a)
                elif i == '':
                    a = 0
                    item_list.append(a)
                else:
                    print("숫자 변환 오류")
            result_list = []
            for i in range(len(item_list)):
                result_list.append(sum_list[i]+item_list[i])
            b = f'■ {date}({status})\n* 예습 : {pre}/10점\n* 복습 : {rev}/10점\n* 수업태도 : {att}/10점\n\n'
            body = body + b
            comments = comments + comm
        footer = ''
        vw = result_list[0]
        va = result_list[1]
        lw = result_list[2]
        la = result_list[3]
        sw = result_list[4]
        sa = result_list[5]
        mw = result_list[6]
        ma = result_list[7]
        if va != 0:
            vr = (va-vw)/va * 100
            vr = round(vr,1)
            v_msg = f'■ Weekly Task\n* 단어(오답/할당량) : -{vw}/{va} (정답률 : {vr}%)'
            footer = footer + v_msg + '\n'
        if la != 0:
            lr = (la-lw)/la * 100
            lr = round(lr,1)
            l_msg = f'* 듣기(오답/할당량) : -{lw}/{la} (정답률 : {lr}%)'
            footer = footer + l_msg + '\n'
        if sa != 0:
            sr = (sa-sw)/sa * 100
            sr = round(sr,1)
            s_msg = f'* 문장암기(오답/할당량) : -{sw}/{sa} (정답률 : {sr}%)'
            footer = footer + s_msg + '\n'
        if ma != 0:
            mr = (ma-mw)/ma * 100
            mr = round(mr,1)
            m_msg = f'* 모의고사(오답/할당량) : -{mw}/{ma} (정답률 : {mr}%)'
            footer = footer + m_msg + '\n'
        if comments != '':
            f_msg = f'※ 특이사항 : {comments}'
            footer = footer + f_msg
        message = init_line + body + footer
        return_list.append([name, message])
    return return_list

def make_xls(msg_list, month_no, week_no):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '전화번호'
    ws['B1'] = "문자내용"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    for i in range(0, len(msg_list)):
        for r in range(0, len(msg_list[i])):
            ws.cell(row = i+2, column = r+1, value = msg_list[i][r])
    now = datetime.datetime.now()
    nowdate = now.strftime('%y%m%d_%H%M%S')
    wb.save(f'주간문자(영어_{month_no}월{week_no}주차_{nowdate}).xlsx')
    print('엑셀 파일 생성 완료.')

def make_txt(msg_list, month_no, week_no):
    cnt = ''
    for i in msg_list:
        cnt = cnt + i[1] + '\n'+ '-'*30+'\n\n'
    now = datetime.datetime.now()
    nowdate = now.strftime('%y%m%d_%H%M%S')
    with open(f'주간문자(영어_{month_no}월{week_no}주차_{nowdate}.txt', 'w', encoding='utf8') as file:
        file.write(cnt)

if __name__ == '__main__':
    os.system('cls')
    print('')
    logo = '\
 #####        #     # ####### #     #  #####  ####### ####### ######\n\
#     #       ##   ## #     # ##    # #     #    #    #       #     # \n\
#     #       # # # # #     # # #   # #          #    #       #     # \n\
#     # ##### #  #  # #     # #  #  #  #####     #    #####   ######  \n\
#   # #       #     # #     # #   # #       #    #    #       #   #   \n\
#    #        #     # #     # #    ## #     #    #    #       #    #  \n\
 #### #       #     # ####### #     #  #####     #    ####### #     #'
    print(logo)
    sharpline = "#" * 69
    opening = f"\n{sharpline}\n\n< 큐몬스터 영어과 주간문자 관리  v.191213 >\n\n{sharpline}\n"
    print(opening)
    # SpreadSheet URL 가져오기
    URL_file = open("./eng_url.txt", 'rt', encoding='UTF8')
    SHEET_URL = URL_file.read()
    URL_file.close()
    print(f"eng_url.txt 파일에서 읽어온 URL : {SHEET_URL}\n")
    month_no = input("몇 월(예 : 12)? ")
    week_no = input("몇 주차(예 : 2)? ")
    print(f"{month_no}월 {week_no}주차 주간 문자 생성 중.")
    list_all = conn_gspred(SHEET_URL, month_no)
    msg_data = write_message(list_all, month_no, week_no)
    input_data = str(input("주간 문자 리스트 생성 완료.\n1 : 주간문자 발송용 엑셀파일 생성.\n2 : 내용 점검용 TXT 파일 생성\n3 : 현재 화면에 내용 출력\n--------------------------\n작업 선택(복수선택 가능 - 예 : 12) : "))
    if "1" in input_data:
        make_xls(msg_data, month_no, week_no)
    elif "2" in input_data:
        make_txt(msg_data, month_no, week_no)
    elif "3" in input_data:
        for i in msg_data:
            print(i[1]+'\n'+ '-'*30+'\n\n')